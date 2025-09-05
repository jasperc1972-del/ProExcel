from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# 建立 Excel 工作簿
wb = Workbook()
ws = wb.active
ws.title = "產品規格書"

# 表格標題
ws.append(["章節", "欄位名稱", "說明", "範例", "內容說明"])

# 產品規格資料
data = [
    ["文件資訊", "文件名稱", "文件標題", "智慧手環產品規格書", ""],
    ["文件資訊", "版本號 / 日期", "修訂版與日期", "V1.0 / 2025-09-05", ""],
    ["文件資訊", "負責人", "文件主責人", "Jasper C", ""],
    ["產品概述", "產品名稱 / 型號", "產品識別", "SmartBand X100", ""],
    ["產品概述", "開發目的", "為什麼要開發", "健康監測，提升生活品質", ""],
    ["產品概述", "目標客群", "產品使用對象", "25–45歲上班族", ""],
    ["功能需求", "核心功能", "必要功能", "心率偵測、步數紀錄、睡眠分析", ""],
    ["功能需求", "次要功能", "加值功能", "音樂控制、訊息提醒", ""],
    ["技術規格", "尺寸 / 重量", "外觀需求", "45mm x 18mm, 25g", ""],
    ["技術規格", "電氣特性", "電壓/電流/功耗", "3.7V / 100mAh / 待機 7天", ""],
    ["技術規格", "通訊方式", "無線協定", "Bluetooth 5.0", ""],
    ["技術規格", "軟體相容性", "系統環境", "iOS 14+ / Android 10+", ""],
    ["品質與標準", "測試項目", "驗證條件", "防水測試、跌落測試", ""],
    ["品質與標準", "認證", "國際標準", "CE, FCC, RoHS", ""],
    ["操作介面", "顯示", "螢幕/介面規格", "1.2\" OLED 解析度 240x240", ""],
    ["環境條件", "工作溫度", "使用溫度範圍", "-10℃ ~ +50℃", ""],
    ["環境條件", "防護等級", "IP 等級", "IP67 防水防塵", ""],
    ["包裝與標示", "包裝尺寸", "外盒大小", "100 x 80 x 50 mm", ""],
    ["包裝與標示", "標籤內容", "條碼/產地", "UPC Code / Made in Taiwan", ""],
    ["開發時程", "里程碑", "重要時間點", "EVT: 2025/10, MP: 2026/03", ""]
]

# 寫入資料
for row in data:
    ws.append(row)

# 建立下拉選單
dv_cert = DataValidation(type="list", formula1='"CE,FCC,RoHS,UL,ISO"', allow_blank=True)
dv_ip = DataValidation(type="list", formula1='"IP65,IP66,IP67,IP68"', allow_blank=True)
dv_milestone = DataValidation(type="list", formula1='"EVT,DVT,PVT,MP"', allow_blank=True)
ws.add_data_validation(dv_cert)
ws.add_data_validation(dv_ip)
ws.add_data_validation(dv_milestone)

# 定義對應內容
cert_content = {
    "CE": "符合歐盟產品安全、環保及健康要求",
    "FCC": "符合美國聯邦通訊委員會規範",
    "RoHS": "符合有害物質限制指令",
    "UL": "符合美國安全認證標準",
    "ISO": "符合國際標準化組織標準"
}

ip_content = {
    "IP65": "防塵 + 防水噴射",
    "IP66": "防塵 + 強水柱",
    "IP67": "防塵 + 1米水下30分鐘",
    "IP68": "防塵 + 長時間水下使用"
}

milestone_content = {
    "EVT": "工程驗證樣機測試",
    "DVT": "設計驗證測試",
    "PVT": "量產驗證測試",
    "MP": "正式量產"
}

# 將下拉選單與內容對應
for i, row in enumerate(data, start=2):
    chapter, field = row[0], row[1]
    col_value = get_column_letter(4)  # 範例欄
    col_content = get_column_letter(5)  # 內容說明欄
    if chapter == "品質與標準" and field == "認證":
        dv_cert.add(f"{col_value}{i}")
        ws[f"{col_content}{i}"] = cert_content.get(row[3].split(",")[0], "")
    if chapter == "環境條件" and field == "防護等級":
        dv_ip.add(f"{col_value}{i}")
        ws[f"{col_content}{i}"] = ip_content.get("IP67", "")
    if chapter == "開發時程" and field == "里程碑":
        dv_milestone.add(f"{col_value}{i}")
        ws[f"{col_content}{i}"] = milestone_content.get("EVT", "")

# 儲存檔案
wb.save("產品規格書_自動內容版-1.xlsx")
print("Excel 已生成：產品規格書_自動內容版-1.xlsx")
